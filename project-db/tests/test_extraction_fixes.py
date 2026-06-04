"""Tests for the 2026-06-04 financial-extraction fixes.

Covers the three root causes the PM hit:
  1. candidate selection silently dropping real docs (923 Rockland = 0 records),
  2. XLSX extraction being unbounded + structureless,
  3. (prompt fixes are validated live, not here).
All deterministic / offline.
"""
from __future__ import annotations

import io

import pytest

from decimal import Decimal

from project_db.ai.financials import (
    _NONTRANSACTIONAL_MODEL_RE,
    _financial_score,
    _select_financial_documents,
    content_is_rollup,
)
from project_db.ai.views import report_project_financials
from project_db.connectors.gdrive.extractors import extract_xlsx
from project_db.db.models import Document, FinancialRecord
from project_db.db.models.docs import DocumentText

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_bytes(sheets: dict[str, list[list]]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for title, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        first = False
        for r in rows:
            ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class TestXlsxExtraction:
    def test_header_and_rows_preserved(self):
        raw = _xlsx_bytes({"Quote": [["Item", "Cost"], ["Tiles", 1000], ["Labour", 2000]]})
        text, method = extract_xlsx(raw)
        assert method == "xlsx-openpyxl"
        assert "### Quote" in text
        assert "Item\tCost" in text          # header row kept
        assert "Tiles\t1000" in text

    def test_huge_sheet_is_capped(self):
        rows = [["col"]] + [[i] for i in range(5000)]
        text, _ = extract_xlsx(_xlsx_bytes({"Big": rows}))
        assert "more row(s) not shown" in text   # truncation noted
        assert len(text) < 100_000               # not the full 5000 rows

    def test_empty_sheet_is_flagged(self):
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.title = "Empty"
        bio = io.BytesIO()
        wb.save(bio)
        text, _ = extract_xlsx(bio.getvalue())
        assert "uncomputed formulas" in text


class TestFinancialScore:
    def test_text_body_scores_when_name_does_not(self):
        # 923 Rockland's "Final SOW.pdf": no keyword in name/folder, money in body.
        score = _financial_score(
            "Final SOW.pdf", "ACTIVE/923 Rockland",
            "Scope of work. Payment schedule: 25% deposit on signing. Total $66,539.",
        )
        assert score > 0

    def test_neutral_doc_still_zero(self):
        assert _financial_score("Notes.pdf", "ACTIVE/X", "meeting notes, no money") == 0

    def test_model_regex_matches_analysis_sheets(self):
        for n in ["Multifamily Acquisition Model v3.xlsx", "Erik_Proforma.xlsx",
                  "Scoring Model.xlsx", "PIPELINE TRACKER.xlsx", "CMHC Data.xlsx"]:
            assert _NONTRANSACTIONAL_MODEL_RE.search(n), n
        # ...but NOT a real cost doc
        assert not _NONTRANSACTIONAL_MODEL_RE.search("Cost Breakdown.xlsx")
        assert not _NONTRANSACTIONAL_MODEL_RE.search("Final SOW.pdf")


class TestCandidateSelection:
    def _doc(self, session, p, name, mime, body):
        d = Document(name=name, url=f"x://{name}", mime_type=mime,
                     project_id=p.canonical_id)
        session.add(d)
        session.flush()
        session.add(DocumentText(document_id=d.canonical_id, extracted_text=body,
                                 extraction_method="t"))
        session.flush()
        return d

    def test_923_rockland_docs_now_selected(self, session, project_factory):
        p = project_factory(name="923 Rockland (3rd Floor)")
        self._doc(session, p, "Final SOW.pdf", "application/pdf",
                  "Scope of work. Payment schedule 25% deposit. Total $66,539.")
        self._doc(session, p, "preliminary quoting file.xlsx", XLSX_MIME,
                  "### Quote\nEstimate # 25008\nTotal Cost\t66539")
        self._doc(session, p, "Multifamily Acquisition Model.xlsx", XLSX_MIME,
                  "### Model\nIRR\t0.18\nNOI\t500000")  # should be skipped
        session.commit()

        cands = _select_financial_documents(
            session, p.canonical_id,
            max_documents=50, per_doc_char_cap=8000, total_char_budget=100_000,
        )
        names = {c.document.name for c in cands}
        assert "Final SOW.pdf" in names                       # was silently dropped before
        assert "preliminary quoting file.xlsx" in names       # "quoting" now matches
        assert "Multifamily Acquisition Model.xlsx" not in names  # model skipped

    def test_contract_pdf_with_no_keyword_still_read(self, session, project_factory):
        p = project_factory(name="Plain Contract Proj")
        self._doc(session, p, "Agreement.pdf", "application/pdf",
                  "This agreement is between the parties for renovation work.")
        session.commit()
        cands = _select_financial_documents(
            session, p.canonical_id,
            max_documents=50, per_doc_char_cap=8000, total_char_budget=100_000,
        )
        assert "Agreement.pdf" in {c.document.name for c in cands}


class TestRollupRecompute:
    """Roll-up status is re-derived at REPORT time (free cleanup, no re-extract)."""

    def test_content_is_rollup(self):
        assert content_is_rollup("Current Project Overview ... Total Project Projection 549241")
        assert content_is_rollup("Pro Forma model: IRR 18%, NOI 500000, cap rate 5%")
        assert not content_is_rollup("Invoice #123 for tile work, total 5000")

    def _seed(self, session, p, name, *, amount, direction, body=""):
        d = Document(name=name, url=f"x://{name}", mime_type=XLSX_MIME,
                     project_id=p.canonical_id)
        session.add(d)
        session.flush()
        if body:
            session.add(DocumentText(document_id=d.canonical_id,
                                     extracted_text=body, extraction_method="t"))
        session.add(FinancialRecord(
            project_id=p.canonical_id, document_id=d.canonical_id,
            direction=direction, record_kind="total", amount=Decimal(str(amount)),
            is_rollup=False))
        session.flush()

    def test_budget_named_doc_excluded_from_totals(self, session, project_factory):
        p = project_factory(name="Budget Proj")
        self._seed(session, p, "C61 revamp budget.xlsx", amount=400000, direction="unknown")
        self._seed(session, p, "Invoice.pdf", amount=5000, direction="contractor_out")
        session.commit()
        rep = report_project_financials(session, str(p.canonical_id))
        assert rep["primary_record_count"] == 1          # budget moved to rollup
        assert rep["rollup_record_count"] == 1
        assert rep["totals"]["contractor_out"] == 5000.0
        assert rep["totals"]["unknown"] == 0.0           # the $400k junk gone from totals

    def test_projection_content_doc_excluded_despite_innocent_name(
        self, session, project_factory
    ):
        p = project_factory(name="Proj Content")
        self._seed(session, p, "Quoting File.xlsx", amount=549241, direction="unknown",
                   body="Current Project Overview\nTotal Project Projection\t549241")
        session.commit()
        rep = report_project_financials(session, str(p.canonical_id))
        assert rep["primary_record_count"] == 0          # content-detected projection
        assert rep["rollup_record_count"] == 1
