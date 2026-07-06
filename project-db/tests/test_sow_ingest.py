"""Real SOW-file ingester: a template SOW workbook -> SowPackage/SowItem rows.

Uses a synthetic in-memory workbook (openpyxl) shaped exactly like the settled
SOW template, so the test is self-contained (no dependency on any real file).
"""

from __future__ import annotations

import io
import uuid

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from project_db.ai.sow_ingest import ingest_sow_workbook
from project_db.db.base import Base
from project_db.db.migrations import ensure_sqlite_schema
from project_db.db.models import Client, Organization, Project, SowItem, SowPackage
from project_db.db.models.work import ProjectStatus

_HEADERS = ["Item_ID", "CSI_Div_Code", "Trade", "Description", "Included", "Material_Spec", "Notes"]
_ROWS = [
    # div 01 = General Requirements => GC overhead, no package
    ["SOW-001", "01", "General Requirements", "Site supervision", "Y", "", "1 supervisor"],
    ["SOW-002", "01", "General Requirements", "Coordinate trades", "Y", "", ""],
    # div 22 plumbing
    ["SOW-003", "22", "Plumbing", "Rough-in", "Y", "PEX-A", ""],
    ["SOW-004", "22", "Plumbing", "Fixtures", "Y", "Moen", ""],
    # div 09 finishes with an explicit EXCLUSION
    ["SOW-005", "09", "Flooring", "LVP throughout", "Y", "6mm", ""],
    ["SOW-006", "09", "Flooring", "Hardwood refinish (by owner)", "N", "", "excluded"],
    # a range division spelled dashless in the sheet -> must canonicalize to 10-12
    ["SOW-007", "1012", "Fixtures", "Kitchen cabinets", "Y", "IKEA", ""],
]


def _make_workbook_bytes(rows=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("SOW_Items")
    ws.append(_HEADERS)
    for r in (rows if rows is not None else _ROWS):
        ws.append(r)
    pc = wb.create_sheet("Parser_Contract")
    pc.append(["Sheet_Name", "Ingest", "Table_Name", "Primary_Key", "Notes"])
    pc.append(["SOW_Items", "Y", "tblSowItems", "Item_ID", "scope"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_engine():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    ensure_sqlite_schema(engine)
    return engine


def _project(session, *, code="2026001"):
    org = Organization(canonical_id=uuid.uuid4(), name="Org")
    client = Client(canonical_id=uuid.uuid4(), name="Client", organization_id=org.canonical_id)
    session.add_all([org, client])
    session.flush()
    p = Project(
        canonical_id=uuid.uuid4(), name="923-927 Rockland", code=code,
        client_id=client.canonical_id, status=ProjectStatus.ACTIVE,
    )
    session.add(p)
    session.flush()
    return p


@pytest.fixture
def session():
    return sessionmaker(bind=_make_engine())()


class TestSowIngest:
    def test_creates_packages_and_items(self, session):
        p = _project(session)
        res = ingest_sow_workbook(session, p, _make_workbook_bytes(), source_name="test.xlsx")
        session.commit()

        assert res.items_created == 7
        assert res.included_items == 6
        assert res.excluded_items == 1
        # packages: 22, 09, 10-12 (NOT 01 -- GC overhead) = 3
        assert res.packages_created == 3
        assert session.query(SowItem).filter_by(project_id=p.canonical_id).count() == 7
        assert session.query(SowPackage).filter_by(project_id=p.canonical_id).count() == 3

    def test_division_01_items_have_no_package(self, session):
        p = _project(session)
        ingest_sow_workbook(session, p, _make_workbook_bytes())
        session.commit()
        gc = session.query(SowItem).filter_by(project_id=p.canonical_id, division_code="01").all()
        assert len(gc) == 2
        assert all(i.package_id is None for i in gc)

    def test_range_division_canonicalized(self, session):
        """Sheet spells '1012'; it must be stored as '10-12' so it matches the
        resolver + budget codes (the 1012-vs-10-12 trap)."""
        p = _project(session)
        ingest_sow_workbook(session, p, _make_workbook_bytes())
        session.commit()
        item = session.query(SowItem).filter_by(
            project_id=p.canonical_id, item_code="SOW-007"
        ).one()
        assert item.division_code == "10-12"
        pkg = session.query(SowPackage).filter_by(
            project_id=p.canonical_id, division_code="10-12"
        ).one()
        assert item.package_id == pkg.canonical_id

    def test_excluded_item_flag(self, session):
        p = _project(session)
        ingest_sow_workbook(session, p, _make_workbook_bytes())
        session.commit()
        excl = session.query(SowItem).filter_by(
            project_id=p.canonical_id, item_code="SOW-006"
        ).one()
        assert excl.included is False

    def test_idempotent_replace(self, session):
        """Re-ingesting replaces prior SOW rows -- no accumulation, no dup codes."""
        p = _project(session)
        ingest_sow_workbook(session, p, _make_workbook_bytes())
        session.commit()
        # A trimmed second version (fewer rows) must fully replace, not append.
        ingest_sow_workbook(session, p, _make_workbook_bytes(rows=_ROWS[:3]))
        session.commit()
        assert session.query(SowItem).filter_by(project_id=p.canonical_id).count() == 3

    def test_missing_sow_items_sheet_raises(self, session):
        p = _project(session)
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="SOW_Items"):
            ingest_sow_workbook(session, p, buf.getvalue())

    def test_accepts_file_path(self, session, tmp_path):
        p = _project(session)
        f = tmp_path / "2026001_SOW_v1.xlsx"
        f.write_bytes(_make_workbook_bytes())
        res = ingest_sow_workbook(session, p, f)
        session.commit()
        assert res.items_created == 7
