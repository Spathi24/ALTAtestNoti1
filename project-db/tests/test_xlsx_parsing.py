"""Slice 3: XlsxParser (openpyxl) -- structure-preserving spreadsheet parsing.

Builds synthetic workbooks with openpyxl to exercise the features real Home
Depot exports lack (formulas, merged cells, title rows above the header,
multiple sheets). Correctness on the 115 real HD files was validated by a
manual in/out sweep (0 failures, 0 anomalies) during development.
"""

from __future__ import annotations

import io
import json

import openpyxl

from project_db.db.models.docs import Document, DocumentText, EvidenceSpan
from project_db.parsing import XlsxParser, get_parser_for, parse_document_content

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _wb_bytes(build) -> bytes:
    wb = openpyxl.Workbook()
    build(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _two_sheet_quote_and_costs() -> bytes:
    def build(wb):
        qs = wb.active
        qs.title = "Quote"
        qs["A1"] = "CLIENT QUOTE - 1455 St Mathieu"
        qs.merge_cells("A1:C1")  # title row above the header
        qs["A2"], qs["B2"], qs["C2"] = "Item", "Qty", "Total"
        qs["A3"], qs["B3"], qs["C3"] = "Window", 2, 1080.45
        qs["A4"], qs["B4"], qs["C4"] = "Door", 1, 559.99
        qs["A5"], qs["C5"] = "TOTAL", "=SUM(C3:C4)"
        qs["C5"].number_format = "$#,##0.00"
        cs = wb.create_sheet("Costs")
        cs.append(["Type", "Company", "Quote"])
        cs.append(["Plumbing", "Brass Mecanique", "43090,33"])
        cs.append(["HVAC", "NouvelAir", "37930,26"])

    return _wb_bytes(build)


# --------------------------------------------------------------------------- #
# Parser (pure)
# --------------------------------------------------------------------------- #


def test_xlsx_can_parse_by_mime_and_extension():
    p = XlsxParser()
    assert p.can_parse(mime=XLSX_MIME, filename=None)
    assert p.can_parse(mime=None, filename="Book.XLSX")
    assert not p.can_parse(mime=None, filename="legacy.xls")  # openpyxl can't read .xls
    assert not p.can_parse(mime="text/csv", filename="x.csv")


def test_router_routes_xlsx():
    assert isinstance(get_parser_for(mime=XLSX_MIME, filename=None), XlsxParser)
    assert isinstance(get_parser_for(mime=None, filename="a.xlsx"), XlsxParser)


def test_xlsx_multisheet_headers_and_samples():
    parsed = XlsxParser().parse(_two_sheet_quote_and_costs(), doc_name="q.xlsx", mime=XLSX_MIME)
    assert parsed.structured["n_sheets"] == 2
    assert parsed.structured["sheet_names"] == ["Quote", "Costs"]
    assert len(parsed.evidence_spans) == 2

    by_sheet = {s.content_json["sheet"]: s for s in parsed.evidence_spans}
    costs = by_sheet["Costs"].content_json
    assert costs["headers"] == ["Type", "Company", "Quote"]
    # Values stay bound to their column -- the 3940 lesson.
    assert costs["rows_sample"][0] == {
        "Type": "Plumbing",
        "Company": "Brass Mecanique",
        "Quote": "43090,33",
    }


def test_xlsx_detects_header_below_title_row():
    span = next(
        s
        for s in XlsxParser()
        .parse(_two_sheet_quote_and_costs(), doc_name="q", mime=XLSX_MIME)
        .evidence_spans
        if s.content_json["sheet"] == "Quote"
    )
    # Row 1 is a merged title; the real header is row 2.
    assert span.locator["header_row"] == 2
    assert span.content_json["headers"] == ["Item", "Qty", "Total"]
    # The title row is not lost -- preserved in the raw grid + rendering.
    assert any("CLIENT QUOTE" in str(c) for row in span.content_json["rows_preview"] for c in row)


def test_xlsx_captures_formula_and_merged_ranges():
    span = next(
        s
        for s in XlsxParser()
        .parse(_two_sheet_quote_and_costs(), doc_name="q", mime=XLSX_MIME)
        .evidence_spans
        if s.content_json["sheet"] == "Quote"
    )
    assert span.content_json["merged_ranges"] == ["A1:C1"]
    cells = span.content_json["cells"]
    assert "C5" in cells  # the SUM total is citeable and distinct from line items
    assert cells["C5"]["formula"] == "=SUM(C3:C4)"
    assert cells["C5"]["number_format"] == "$#,##0.00"


def test_xlsx_empty_sheet_is_handled():
    def build(wb):
        wb.active.title = "Blank"  # no data

    parsed = XlsxParser().parse(_wb_bytes(build), doc_name="e.xlsx", mime=XLSX_MIME)
    assert parsed.structured["n_sheets"] == 1
    assert parsed.evidence_spans == []  # no table_region for an empty sheet


# --------------------------------------------------------------------------- #
# Persisted spine via the service
# --------------------------------------------------------------------------- #


def test_parse_document_content_xlsx_full_spine(session):
    doc = Document(name="quote.xlsx", url="https://drive/quote.xlsx", mime_type=XLSX_MIME)
    session.add(doc)
    session.commit()

    parse = parse_document_content(session, document=doc, content=_two_sheet_quote_and_costs())
    session.commit()

    assert parse.status == "success" and parse.parser_name == "xlsx"
    assert json.loads(parse.structured_json)["n_sheets"] == 2
    assert session.query(EvidenceSpan).filter_by(parse_id=parse.id).count() == 2

    dt = session.query(DocumentText).filter_by(document_id=doc.canonical_id).one()
    assert dt.extraction_method == "xlsx/1"
    assert "## Sheet: Quote" in dt.extracted_text and "## Sheet: Costs" in dt.extracted_text
