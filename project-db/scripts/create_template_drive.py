"""Create mock template Drive for project 2026001 — Rockland.

Generates:
  docs/templates/mock_drive/
    2026001 — Rockland/
      JOBCOST/  SOW/  SOW/packages/  quotes/  green-sheet/  POs/  budget/

Run from repo root:
  python project-db/scripts/create_template_drive.py

Column-shape rules:
  QUOTE / PKG  -> grid-readable: Material Amount / Labour Amount / Total Amount
                  (deterministic grid parser reads these to the penny)
  SOW          -> scope grid: Item_ID / CSI_Div_Code / Description / Included / Material_Spec
  JOBCOST      -> Cost_Ledger shape (mirrors JOB_COST_TEMPLATE_structured.xlsx)
  BUDGET       -> Budget_Lines flat table
  PO           -> PO_Header key-value + PO_Lines with M/L/T
  GREENSHEET   -> display only, all sheets Ingest=N

Every file carries a Parser_Contract sheet so ALTA knows which sheets to ingest.
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

# Paths
_HERE = Path(__file__).parent
REPO_ROOT = _HERE.parent.parent
MOCK_DRIVE = REPO_ROOT / "docs" / "templates" / "mock_drive"
PROJECT_DIR = MOCK_DRIVE / "2026001 — Rockland"

PC = "2026001"  # project code

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")  # light blue
META_FILL = PatternFill("solid", fgColor="F2F2F2")    # light gray
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow

# Real Rockland CSI divisions (from DB query, excl. 01-general / 99-unclassified)
TRADES = [
    ("02", "Demolition"),
    ("03", "Concrete"),
    ("05", "Structural"),
    ("06", "Carpentry"),
    ("07", "Roofing"),
    ("08", "DoorsWindows"),
    ("09", "Finishes"),
    ("1012", "Fixtures"),
    ("22", "Plumbing"),
    ("23", "HVAC"),
    ("26", "Electrical"),
]

# Rockland SOW items (reflects real line-item data from the DB)
SOW_ITEMS = [
    # (Item_ID, CSI_Div_Code, Trade, Description, Included, Material_Spec, Notes)
    ("SOW-001", "01", "General Requirements", "Site supervision and coordination", "Y", "", ""),
    ("SOW-002", "01", "General Requirements", "Overhead, profit (15%), and contingency (3%)", "Y", "", ""),
    ("SOW-003", "01", "General Requirements", "Construction material deliveries", "Y", "", ""),
    ("SOW-004", "02", "Demolition", "Demo existing bathroom fixtures and tile", "Y", "", "All debris removed"),
    ("SOW-005", "02", "Demolition", "Remove existing flooring throughout unit", "Y", "", ""),
    ("SOW-006", "02", "Demolition", "Waste disposal / debris removal", "Y", "", ""),
    ("SOW-007", "03", "Concrete", "Concrete patching at subfloor penetrations", "Y", "", "If required"),
    ("SOW-008", "05", "Structural", "Structural engineer review for load-bearing elements", "Y", "", ""),
    ("SOW-009", "05", "Structural", "Steel beam / lintel supply and install if required", "Y", "", "Conditional"),
    ("SOW-010", "06", "Carpentry", "Install new door frames and trim throughout", "Y", "Finger-jointed pine", ""),
    ("SOW-011", "06", "Carpentry", "Install baseboards and crown moulding", "Y", "MDF painted", ""),
    ("SOW-012", "06", "Carpentry", "Blocking and backing for fixtures", "Y", "", ""),
    ("SOW-013", "07", "Roofing", "Roof replacement with ventilation system", "Y", "Fibreglass shingles, 25-yr", ""),
    ("SOW-014", "07", "Roofing", "Disposal of existing roofing material", "Y", "", ""),
    ("SOW-015", "08", "DoorsWindows", "Supply and install interior doors", "Y", "Hollow-core, pre-hung", ""),
    ("SOW-016", "08", "DoorsWindows", "Supply and install sliding closet doors", "Y", "Mirrored bypass", ""),
    ("SOW-017", "08", "DoorsWindows", "New opening for sliding doors (framing)", "Y", "", ""),
    ("SOW-018", "09", "Finishes", "Drywall and taping throughout", "Y", "5/8\" Type X", ""),
    ("SOW-019", "09", "Finishes", "Tile floor in bathroom incl. waterproofing membrane", "Y", "12x24 porcelain, client-selected", ""),
    ("SOW-020", "09", "Finishes", "Painting throughout (2 coats)", "Y", "Benjamin Moore Chantilly Lace", ""),
    ("SOW-021", "09", "Finishes", "LVP flooring throughout living areas", "Y", "Lifeproof 6mm, client-selected", ""),
    ("SOW-022", "1012", "Fixtures", "Supply and install kitchen cabinets", "Y", "IKEA SEKTION white", ""),
    ("SOW-023", "1012", "Fixtures", "Supply and install bathroom vanity", "Y", "36\" single sink, white", ""),
    ("SOW-024", "1012", "Fixtures", "Supply and install countertop", "Y", "Laminate standard", ""),
    ("SOW-025", "22", "Plumbing", "Rough-in plumbing (drain, supply, vent)", "Y", "PEX-A", ""),
    ("SOW-026", "22", "Plumbing", "Supply and install plumbing fixtures (toilet, sink, shower)", "Y", "Moen Adler series", ""),
    ("SOW-027", "22", "Plumbing", "Hot water heater replacement", "Y", "40-gal electric", ""),
    ("SOW-028", "23", "HVAC", "Supply and install bathroom exhaust fan", "Y", "90 CFM, ENERGY STAR", ""),
    ("SOW-029", "23", "HVAC", "Electric baseboard heaters throughout", "Y", "Stelpro 1500W per room", ""),
    ("SOW-030", "26", "Electrical", "Electrical rough-in (circuits, panel upgrade)", "Y", "200A panel", ""),
    ("SOW-031", "26", "Electrical", "Supply and install light fixtures throughout", "Y", "LED, client-selected", ""),
]


# ── helpers ───────────────────────────────────────────────────────────────────

def _wb() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def _sheet(
    wb: openpyxl.Workbook,
    name: str,
    headers: list[str],
    rows: list[list],
    *,
    meta: bool = False,
) -> None:
    ws = wb.create_sheet(name)
    fill = META_FILL if meta else HEADER_FILL
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = fill
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    ws.freeze_panes = "A2"
    for c, h in enumerate(headers, 1):
        col_vals = [str(row[c - 1]) for row in rows if c - 1 < len(row)]
        width = max(len(str(h)), max((len(v) for v in col_vals), default=0)) + 2
        ws.column_dimensions[get_column_letter(c)].width = min(width, 45)


def _pc_sheet(wb: openpyxl.Workbook, entries: list[tuple]) -> None:
    """Add Parser_Contract sheet. entries: (Sheet, Ingest, Table, PK, Notes)."""
    _sheet(
        wb, "Parser_Contract",
        ["Sheet_Name", "Ingest", "Table_Name", "Primary_Key", "Notes"],
        [list(e) for e in entries],
        meta=True,
    )


def _save(wb: openpyxl.Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  OK  {path.relative_to(REPO_ROOT)}")


# ── file builders ─────────────────────────────────────────────────────────────

def build_sow(path: Path) -> None:
    wb = _wb()

    # README
    ws = wb.create_sheet("README")
    ws["A1"] = "SOW Template — 2026001 Rockland"
    ws["A1"].font = Font(bold=True, size=12)
    for r, (k, v) in enumerate([
        ("Purpose", "Defines the contract boundary for this project."),
        ("Included=Y", "Item is in the base contract price."),
        ("Included=N", "Excluded — any work added later requires a Change Order."),
        ("Rule", "Do not silently absorb out-of-scope work. Track every change."),
    ], 3):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)

    _sheet(wb, "SOW_Items",
           ["Item_ID", "CSI_Div_Code", "Trade", "Description", "Included", "Material_Spec", "Notes"],
           [list(i) for i in SOW_ITEMS])

    _pc_sheet(wb, [
        ("README",   "N", "—", "—", "Human instructions — do not ingest"),
        ("SOW_Items","Y", "tblSowItems", "Item_ID",
         "Contract scope. Included=Y/N defines boundary. Ingest during Phase 3 (SowItem model)."),
        ("Parser_Contract", "N", "—", "—", "Parser instructions (this sheet)"),
    ])
    _save(wb, path)


def build_package(path: Path, csi: str, trade: str) -> None:
    wb = _wb()

    # Filter SOW items for this trade
    items = [
        [i[0], i[0], i[3], i[5], "", "", i[6]]
        for i in SOW_ITEMS if i[1] == csi
    ]
    if not items:
        items = [
            [f"{csi}-001", f"{csi}-001", f"{trade} scope item 1", "", "", "", ""],
            [f"{csi}-002", f"{csi}-002", f"{trade} scope item 2", "", "", "", ""],
        ]

    _sheet(wb, "Package_Lines",
           ["Item_ID", "SOW_Item_Ref", "Description", "Material_Spec", "Qty", "Unit", "Notes"],
           items)

    _pc_sheet(wb, [
        ("Package_Lines", "Y", "tblPackageLines", "Item_ID",
         "Trade scope sent to subcontractors for quoting. No money columns — amounts come from QUOTE files."),
        ("Parser_Contract", "N", "—", "—", "Parser instructions"),
    ])
    _save(wb, path)


def build_quote(
    path: Path,
    csi: str,
    trade: str,
    vendor: str,
    lines: list[tuple],  # (desc, csi_hint, mat, lab, tot, coverage, mat_incl, excl, notes)
) -> None:
    """Column order is deliberate:
    Description | Masterformat | Material Amount | Labour Amount | Total Amount | ...extras
    This ensures _map_columns picks Material Amount (not Materials_Included) as the
    material column, and Description (not Item_ID) as the description column.

    Structure:
      Row 1 = header
      Row 2 = section-total row: Description=trade, Masterformat=CSI, Total Amount=sum
              (no material/labour → grid parser sets current division)
      Rows 3+ = line items: Description=item desc, mat+labour filled, total empty
      Last row = Pre-Tax Total: Description EMPTY (parser routes to grand_total branch),
                 Total Amount = sum, Item_ID col = "Pre-Tax Total" label
    """
    wb = _wb()

    # Headers: money cols immediately after Description/Masterformat to avoid shadowing.
    # "Mat_Incl" (not "Materials_Included") avoids triggering "material" substring match.
    HEADERS = [
        "Description", "Masterformat",
        "Material Amount", "Labour Amount", "Total Amount",
        "Item_ID", "Coverage_Y_N", "Mat_Incl", "Exclusions", "Notes",
    ]
    pretax = sum(ln[4] for ln in lines)

    # Section-total row (sets current division; no material/labour values)
    section_row = [trade, csi, "", "", pretax, "", "", "", "", ""]
    # Line-item rows (material + labour filled; total empty)
    item_rows = []
    for idx, (desc, csi_hint, mat, lab, tot, cov, mat_incl, excl, notes) in enumerate(lines, 1):
        item_rows.append([desc, csi_hint, mat, lab, "", f"QI-{idx:03d}", cov, mat_incl, excl, notes])

    all_rows = [section_row] + item_rows
    _sheet(wb, "Quote_Lines", HEADERS, all_rows)

    # Pre-Tax Total row: Description EMPTY so parser routes to grand_total branch,
    # scans full row for "Pre-Tax Total" label, captures Total Amount as grand_total.
    ws = wb["Quote_Lines"]
    last = len(all_rows) + 2
    ws.cell(last, 1, "")            # Description = empty (triggers grand_total path)
    ws.cell(last, 5, pretax)        # Total Amount col (index 4 → col 5 in 1-based)
    ws.cell(last, 6, "Pre-Tax Total")  # Item_ID col carries the label (non-description col)
    ws.cell(last, 6).font = Font(bold=True)
    ws.cell(last, 5).font = Font(bold=True)
    for c in range(1, 11):
        ws.cell(last, c).fill = TOTAL_FILL

    _pc_sheet(wb, [
        ("Quote_Lines", "Y", "tblQuoteLines", "Item_ID",
         "Subcontractor quote. Grid-readable: Material Amount / Labour Amount / Total Amount. "
         "Row 2 = section total (sets CSI division). Line items follow. "
         "Pre-Tax Total row has empty Description so grand_total is captured."),
        ("Parser_Contract", "N", "—", "—", "Parser instructions"),
    ])
    _save(wb, path)


def build_greensheet(path: Path) -> None:
    wb = _wb()

    rows = [
        ["01",   "General Requirements", 9000,  9000,  "",    "",    9000,  9000,  9000,  9000, 0,    "awarded"],
        ["02",   "Demolition",           5000,  4800,  5200,  "",    4800,  4800,  4800,  4200, 600,  "awarded"],
        ["03",   "Concrete",             2000,  1800,  "",    "",    1800,  1800,  1800,  1800, 0,    "awarded"],
        ["05",   "Structural",           3500,  3200,  "",    "",    3200,  3200,  3200,  3200, 0,    "awarded"],
        ["06",   "Carpentry",            8000,  7800,  8200,  "",    7800,  7800,  7800,  0,    7800, "committed"],
        ["07",   "Roofing",              9300,  9300,  "",    "",    9300,  9300,  9300,  9300, 0,    "awarded"],
        ["08",   "DoorsWindows",         6000,  5800,  6100,  "",    5800,  5800,  5800,  0,    5800, "committed"],
        ["09",   "Finishes",             22000, 21500, 22800, 23000, 21500, 21500, 21500, 0,    21500,"committed"],
        ["1012", "Fixtures",             9000,  8800,  9200,  "",    8800,  8800,  8800,  0,    8800, "pending"],
        ["22",   "Plumbing",             7000,  6800,  7200,  "",    6800,  6800,  6800,  0,    6800, "selected"],
        ["23",   "HVAC",                 4500,  4200,  4800,  "",    4200,  4200,  4200,  0,    4200, "pending"],
        ["26",   "Electrical",           8000,  7700,  8300,  "",    7700,  7700,  7700,  0,    7700, "pending"],
    ]
    _sheet(wb, "Green_Sheet",
           ["CSI_Div_Code", "Trade", "Alta_Cost", "Quote_1", "Quote_2", "Quote_3",
            "Selected_Amount", "PO_Amount", "Committed", "Actual", "Variance", "Status"],
           rows)

    _pc_sheet(wb, [
        ("Green_Sheet", "N", "—", "—",
         "Display/summary only. Computed by ALTA green-sheet report. Do NOT ingest as source data."),
        ("Parser_Contract", "N", "—", "—", "Parser instructions"),
    ])
    _save(wb, path)


def build_po(path: Path) -> None:
    wb = _wb()

    _sheet(wb, "PO_Header",
           ["Field", "Value"],
           [
               ["PO_Number",             "2026001-001"],
               ["Project_Code",          "2026001"],
               ["Project_Display",       "2026001 — Rockland"],
               ["Vendor",                "Plombert Inc."],
               ["Trade_Type",            "Subcontractor"],
               ["Purchase_Type",         "vendor"],
               ["CSI_Div_Code",          "22"],
               ["Trade",                 "Plumbing"],
               ["Contract_Amount",       6800.00],
               ["Currency",              "CAD"],
               ["Tax_Rate",              0.15],
               ["Contract_Amount_Incl_Tax", 7820.00],
               ["Status",                "awarded"],
               ["Issue_Date",            "2026-07-01"],
               ["Payment_Terms",         "Net 30"],
               ["Notes", "Covers SOW-025, SOW-026, SOW-027 per 2026001_PKG_22-Plumbing.xlsx"],
           ])

    _sheet(wb, "PO_Lines",
           ["Line_ID", "SOW_Item_Ref", "Description",
            "Material Amount", "Labour Amount", "Total Amount", "Notes"],
           [
               ["PO-001-01", "SOW-025", "Rough-in plumbing (drain, supply, vent)",  800.00, 2400.00, 3200.00, "PEX-A"],
               ["PO-001-02", "SOW-026", "Supply and install plumbing fixtures",     1600.00, 1200.00, 2800.00, "Moen Adler"],
               ["PO-001-03", "SOW-027", "Hot water heater replacement",              500.00,  300.00,  800.00, "40-gal electric"],
           ])

    _pc_sheet(wb, [
        ("PO_Header", "Y", "tblPOHeader", "Field",
         "Key-value metadata for this PO. Read by field Code."),
        ("PO_Lines", "Y", "tblPOLines", "Line_ID",
         "Line items. Grid-readable: Material Amount / Labour Amount / Total Amount."),
        ("Parser_Contract", "N", "—", "—", "Parser instructions"),
    ])
    _save(wb, path)


def build_budget(path: Path) -> None:
    wb = _wb()

    # Client_Price = Budget_Amount x Line_Markup_Factor, then global 1.15 applied at report level.
    # We store the pre-global client price here for per-line transparency.
    rows = [
        ["BUD-001", "01",   "General Requirements", 9000.00,  1.00, 9000.00],
        ["BUD-002", "02",   "Demolition",           4800.00,  1.00, 4800.00],
        ["BUD-003", "03",   "Concrete",             1800.00,  1.00, 1800.00],
        ["BUD-004", "05",   "Structural",           3200.00,  1.00, 3200.00],
        ["BUD-005", "06",   "Carpentry",            7800.00,  1.05, 8190.00],
        ["BUD-006", "07",   "Roofing",              9300.00,  1.00, 9300.00],
        ["BUD-007", "08",   "DoorsWindows",         5800.00,  1.00, 5800.00],
        ["BUD-008", "09",   "Finishes",            21500.00,  1.10, 23650.00],
        ["BUD-009", "1012", "Fixtures",             8800.00,  1.00, 8800.00],
        ["BUD-010", "22",   "Plumbing",             6800.00,  1.00, 6800.00],
        ["BUD-011", "23",   "HVAC",                 4200.00,  1.00, 4200.00],
        ["BUD-012", "26",   "Electrical",           7700.00,  1.00, 7700.00],
    ]
    _sheet(wb, "Budget_Lines",
           ["Line_ID", "CSI_Div_Code", "Trade",
            "Budget_Amount", "Line_Markup_Factor", "Client_Price_Pre_Global"],
           rows)

    # Totals
    ws = wb["Budget_Lines"]
    last = len(rows) + 2
    ws.cell(last, 1, "TOTAL").font = Font(bold=True)
    ws.cell(last, 4, sum(r[3] for r in rows)).font = Font(bold=True)
    ws.cell(last, 6, sum(r[5] for r in rows)).font = Font(bold=True)
    for c in range(1, 7):
        ws.cell(last, c).fill = TOTAL_FILL

    # Notes
    ws2 = wb.create_sheet("README")
    ws2["A1"] = "Budget Snapshot — 2026001 Rockland"
    ws2["A1"].font = Font(bold=True, size=12)
    for r, (k, v) in enumerate([
        ("Client_Price_Pre_Global", "= Budget_Amount x Line_Markup_Factor (per-line inflation only)."),
        ("Final Client Price", "= SUM(Client_Price_Pre_Global) x 1.15 (global 15% markup applied at subtotal)."),
        ("Rule", "This snapshot is frozen. Do not edit after owner approval. Create v2 for revisions."),
    ], 3):
        ws2.cell(r, 1, k).font = Font(bold=True)
        ws2.cell(r, 2, v)

    _pc_sheet(wb, [
        ("README",       "N", "—", "—", "Human instructions — do not ingest"),
        ("Budget_Lines", "Y", "tblBudgetLines", "Line_ID",
         "Frozen per-trade budget. Client_Price_Pre_Global = Budget_Amount x Line_Markup_Factor. "
         "Global 15% applied at report level to get final client price."),
        ("Parser_Contract", "N", "—", "—", "Parser instructions"),
    ])
    _save(wb, path)


def build_jobcost(path: Path) -> None:
    """Mirrors JOB_COST_TEMPLATE_structured.xlsx shape."""
    wb = _wb()

    # README
    ws = wb.create_sheet("README")
    ws["A1"] = f"JOBCOST — {PC} Rockland"
    ws["A1"].font = Font(bold=True, size=12)
    for r, (k, v) in enumerate([
        ("Purpose", "Canonical job-cost ledger: actuals (Cost_Ledger) + forecast (Scope_Budget) + change orders + order quantities."),
        ("Main input tables", "tblCostLedger, tblScopeBudget, tblChangeOrders, tblOrderQuantities."),
        ("Summary logic", "Dashboard uses SUMIFS by Cost_Class, Cost_Code, and Include flags."),
        ("Append rule", "Append rows inside named tables. Do not build summaries from fixed row ranges."),
        ("Parser rule", "Only ingest sheets listed Ingest=Y in Parser_Contract."),
    ], 3):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)

    # Project_Setup
    _sheet(wb, "Project_Setup",
           ["Code", "Value", "Units", "Description"],
           [
               ["Project_Name",          "2026001 — Rockland", "text", "Project display name"],
               ["Project_Code",          "2026001",                  "text", "Canonical 7-digit project code"],
               ["Legacy_Job_Number",     "923",                      "text", "Previous short reference"],
               ["Tax_Rate",              0.15,                       "rate", "QC: TPS+TVQ combined"],
               ["Global_Markup_Rate",    0.15,                       "rate", "Applied to inflated subtotal -> client price"],
               ["Total_Units",           1,                          "units","Rockland pilot = 1 unit (single renovation)"],
           ])

    # Cost_Ledger
    _sheet(wb, "Cost_Ledger",
           ["Entry_ID", "Date", "Source", "Cost_Type", "Cost_Class", "Trade", "Cost_Code",
            "Scope_Area", "Item_Description", "Supplier_Subcontractor", "SOW_Item_Ref",
            "Manual_Amount_Incl_Tax", "Tax_Included", "Tax_Rate",
            "Amount_Incl_Tax", "Amount_Excl_Tax",
            "Include_In_Actual", "Include_In_Budget", "Notes"],
           [
               ["MAT-0001", "2026-07-05", "Invoice", "Actual", "Material", "Finishes", "09",
                "Unit 923", "LVP Flooring (Lifeproof 6mm)", "Home Depot", "SOW-021",
                1850.00, "Yes", 0.15, 1850.00, 1608.70, "Yes", "No", "Receipt on file"],
               ["LAB-0001", "2026-07-10", "Invoice", "Actual", "Subcontractor", "Plumbing", "22",
                "Unit 923", "Rough-in plumbing", "Plombert Inc.", "SOW-025",
                3200.00, "Yes", 0.15, 3200.00, 2782.61, "Yes", "No", "PO 2026001-001"],
               ["MAT-0002", "", "Committed", "Committed", "Material", "Finishes", "09",
                "Unit 923", "Porcelain tile 12x24", "Mondial", "SOW-019",
                2400.00, "Yes", 0.15, 2400.00, 2086.96, "No", "Yes", "Order pending"],
           ])

    # Scope_Budget
    _sheet(wb, "Scope_Budget",
           ["Line_ID", "Line_Role", "Source_Type", "Cost_Class", "Trade", "Cost_Code",
            "Scope_Area", "Item", "Unit_Price", "Quantity_Per_Affected_Unit", "Affected_Units",
            "Tax_Included", "Tax_Rate", "Amount_Incl_Tax", "Amount_Excl_Tax",
            "Include_Current_Forecast", "Include_Quote_Total", "Include_Target_Budget", "Notes"],
           [
               ["BUD-0001", "Quote_Total", "Quote", "Subcontractor", "Plumbing", "22",
                "Unit 923", "Plumbing (Plombert Inc. selected quote)", 6800.00, 1, 1,
                "Yes", 0.15, 7820.00, 6800.00, "Yes", "Yes", "Yes",
                "Source: 2026001_QUOTE_22-Plumbing_PlombertInc_selected.xlsx"],
               ["BUD-0002", "Allowance", "Current_Forecast", "Material", "Finishes", "09",
                "Unit 923", "Tile allowance (12x24 porcelain)", 3500.00, 1, 1,
                "Yes", 0.15, 4025.00, 3500.00, "Yes", "No", "Yes", "Client-selected, pending invoice"],
               ["BUD-0003", "Quote_Total", "Quote", "Subcontractor", "Finishes", "09",
                "Unit 923", "Finishes (ABC Tile pending quote)", 21500.00, 1, 1,
                "Yes", 0.15, 24725.00, 21500.00, "Yes", "Yes", "Yes",
                "Source: 2026001_QUOTE_09-Finishes_ABCTile_pending.xlsx"],
           ])

    # Change_Orders
    _sheet(wb, "Change_Orders",
           ["CO_ID", "Item", "Description", "Cost_Before_OHP", "OHP_Rate",
            "Total_With_OHP", "Status", "Billable", "Reason", "SOW_Item_Ref"],
           [
               ["CO-001", "Extra demo", "Additional demo for hidden rot in subfloor",
                800.00, 0.15, 920.00, "Pending", "Yes", "Unforeseen condition", ""],
           ])

    # Order_Quantities
    _sheet(wb, "Order_Quantities",
           ["Order_ID", "Cost_Class", "Cost_Code", "Item",
            "Qty_Per_Pack", "Number_Of_Packs", "Total_Units",
            "Order_Status", "Supplier", "Notes"],
           [
               ["ORD-0001", "Material", "09", "Porcelain tile 12x24 (box/10 sqft)",
                10, 28, 280, "Open", "Mondial", "Bathroom + hallway"],
               ["ORD-0002", "Material", "09", "LVP flooring (box/20 sqft)",
                20, 18, 360, "Ordered", "Home Depot", "Living + bedroom"],
               ["ORD-0003", "Material", "22", "PEX-A 1/2\" (roll/100ft)",
                100, 2, 200, "Ordered", "Home Depot", ""],
           ])

    # Lists (validation reference)
    _sheet(wb, "Lists",
           ["Cost_Class", "Cost_Type", "Order_Status", "CO_Status", "Quote_Status"],
           [
               ["Material",     "Actual",     "Open",      "Pending",      "pending"],
               ["Labour",       "Committed",  "Ordered",   "Paid",         "recommended"],
               ["Subcontractor","Forecast",   "Received",  "Not Executed", "selected"],
               ["Equipment",    "Quote",      "Cancelled", "Cancelled",    "rejected"],
               ["",             "",           "",          "",             "awarded"],
           ],
           meta=True)

    _pc_sheet(wb, [
        ("README",         "N", "—", "—", "Human instructions — do not ingest"),
        ("Project_Setup",  "Y", "tblProjectSettings", "Code",
         "Controlled project variables. Read by Code field."),
        ("Cost_Ledger",    "Y", "tblCostLedger", "Entry_ID",
         "Canonical actuals. Include_In_Actual=Yes -> actuals; Include_In_Budget=Yes -> budget."),
        ("Scope_Budget",   "Y", "tblScopeBudget", "Line_ID",
         "Allowances, quote totals, manual overrides. Inclusion flags control forecast/quote/budget."),
        ("Change_Orders",  "Y", "tblChangeOrders", "CO_ID",
         "Change orders. Status=Paid rows count toward actuals."),
        ("Order_Quantities","Y","tblOrderQuantities", "Order_ID",
         "Materials on order. Order_Status tracks receipt."),
        ("Lists",          "N", "—", "—", "Dropdown validation — do not ingest"),
        ("Parser_Contract","N", "—", "—", "Parser instructions (this sheet)"),
    ])
    _save(wb, path)


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    sys.path.insert(0, str(_HERE.parent / "src"))
    try:
        from project_db.cli import force_utf8_output
        force_utf8_output()
    except Exception:
        pass  # non-fatal if run outside the package

    print(f"Building mock template Drive -> docs/templates/mock_drive/\n")

    build_sow(PROJECT_DIR / "SOW" / f"{PC}_SOW_v1.xlsx")

    for csi, trade in TRADES:
        build_package(PROJECT_DIR / "SOW" / "packages" / f"{PC}_PKG_{csi}-{trade}.xlsx", csi, trade)

    build_quote(
        PROJECT_DIR / "quotes" / f"{PC}_QUOTE_22-Plumbing_PlombertInc_pending.xlsx",
        "22", "Plumbing", "Plombert Inc.",
        [
            ("Rough-in plumbing (drain, supply, vent)",  "22", 800.00,  2400.00, 3200.00, "Y","N","", "PEX-A; per SOW-025"),
            ("Supply and install plumbing fixtures",     "22", 1600.00, 1200.00, 2800.00, "Y","Y","", "Moen Adler; per SOW-026"),
            ("Hot water heater replacement",             "22",  500.00,  300.00,  800.00, "Y","Y","", "40-gal electric; per SOW-027"),
        ],
    )
    build_quote(
        PROJECT_DIR / "quotes" / f"{PC}_QUOTE_09-Finishes_ABCTile_pending.xlsx",
        "09", "Finishes", "ABC Tile",
        [
            ("Bathroom tile supply and install",         "09", 4200.00, 3800.00,  8000.00, "Y","Y","", "12x24 porcelain; incl. waterproofing"),
            ("Drywall and taping throughout",            "09", 1800.00, 4200.00,  6000.00, "Y","N","", "5/8\" Type X"),
            ("LVP flooring supply and install",         "09", 2800.00, 1700.00,  4500.00, "Y","Y","", "Lifeproof 6mm"),
            ("Painting 2 coats throughout",             "09", 1200.00, 1800.00,  3000.00, "Y","N","", "BM Chantilly Lace"),
        ],
    )

    build_greensheet(PROJECT_DIR / "green-sheet" / f"{PC}_GREENSHEET.xlsx")
    build_po(PROJECT_DIR / "POs" / f"{PC}-001_PO_22-Plumbing_PlombertInc_awarded.xlsx")
    build_budget(PROJECT_DIR / "budget" / f"{PC}_BUDGET_v1.xlsx")
    build_jobcost(PROJECT_DIR / "JOBCOST" / f"{PC}_JOBCOST.xlsx")

    print(f"\nAll files written. Run verify step next.")


if __name__ == "__main__":
    main()
