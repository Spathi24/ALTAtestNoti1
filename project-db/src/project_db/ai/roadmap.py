"""Roadmap import / persistence helpers.

Reads ``docs/Project Roadmap.xlsx`` (the canonical design-phase
workflow) and persists rows into the ``roadmap_task`` table.  Pure
functions only -- ``parse_roadmap_xlsx`` is a no-DB-touch parser that
returns plain dicts; ``import_roadmap_rows`` is the DB writer.  Keeping
them separate makes the parser easy to unit-test against fixture files
without dragging a session through every assertion.

Used by:
  - the CLI ``project_db import-roadmap`` command
  - eventually the AI layer's prompt-injection helper (Layer 2),
    which calls ``list_roadmap_tasks(session)`` to build the prompt
    block
  - eventually the gap-finder (Layer 3)
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from project_db.db.models import RoadmapActor, RoadmapPhase, RoadmapTask


# Sentinel values that show up in the xlsx as "blank" but read as
# strings via openpyxl / pandas.  Treat them as empty.
_BLANK_PHASE_TOKENS = frozenset({"", "nan", "none", "null"})


def _looks_blank(value: Any) -> bool:
    """True for None, NaN, or empty / sentinel strings."""
    if value is None:
        return True
    if isinstance(value, float):
        # NaN is the only float that doesn't equal itself.
        return value != value
    if isinstance(value, str):
        return value.strip().lower() in _BLANK_PHASE_TOKENS
    return False


def _split_sub_tasks(raw: Any) -> list[str] | None:
    """Parse the xlsx's "Sub-tasks" column into a clean list[str].

    The xlsx stores bullets one-per-line, each starting with "-".
    Lines without "-" still count if they are non-blank.  Whitespace
    is normalized.  Returns None when the cell is blank, NOT an
    empty list (the two are semantically different: "no sub-tasks"
    vs "the editor explicitly cleared them").
    """
    if _looks_blank(raw):
        return None
    text = str(raw).strip()
    if not text:
        return None
    items: list[str] = []
    for line in text.split("\n"):
        clean = line.strip()
        if not clean:
            continue
        # Strip leading bullets / dashes.
        clean = clean.lstrip("-").lstrip("•").lstrip("*").strip()
        if clean:
            items.append(clean)
    return items or None


def parse_roadmap_xlsx(path: str | Path) -> list[dict[str, Any]]:
    """Parse the roadmap spreadsheet into a list of dicts.

    Pure function -- no DB touch.  Returns one dict per non-blank row:

        {
            "phase": RoadmapPhase,
            "ordinal": int,            # 1-based, ordered within phase
            "task_name": str,
            "sub_tasks": list[str] | None,
            "notes": str | None,
        }

    Blank rows (xlsx separators between phases) are dropped.  Rows
    with an unrecognized phase string raise ``ValueError`` -- it's
    cheaper to fail loudly here than to silently lose data.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to parse the roadmap xlsx.  "
            "Install with: pip install openpyxl"
        ) from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if header is None:
        return []
    # Find columns by header name -- the xlsx editor may shuffle columns,
    # but the names ("Design Phase", "Task", "Notes", "Sub-tasks") are
    # stable.
    lookup = {
        (h or "").strip().lower(): i for i, h in enumerate(header)
    }
    try:
        col_phase = lookup["design phase"]
        col_task = lookup["task"]
    except KeyError as exc:
        raise ValueError(
            f"roadmap xlsx is missing required column {exc.args[0]!r}.  "
            f"Headers found: {[h for h in header]}"
        ) from exc
    col_notes = lookup.get("notes")
    col_sub = lookup.get("sub-tasks")

    out: list[dict[str, Any]] = []
    # Ordinals reset per phase so each phase reads as 1, 2, 3, ...
    ordinals: dict[RoadmapPhase, int] = {p: 0 for p in RoadmapPhase}

    def _safe_get(row: tuple, idx: int | None):
        """Bounds-safe row access.

        openpyxl's read-only iter_rows can return tuples shorter than
        the header row when trailing cells are empty.  Treat
        out-of-range as a missing value, NOT as an error.
        """
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in rows_iter:
        if row is None:
            continue
        phase_raw = _safe_get(row, col_phase)
        task_raw = _safe_get(row, col_task)
        if _looks_blank(phase_raw) or _looks_blank(task_raw):
            # Editorial blank row (phase separator) -- skip.
            continue
        phase_str = str(phase_raw).strip().upper()
        try:
            phase = RoadmapPhase(phase_str)
        except ValueError as exc:
            raise ValueError(
                f"unknown phase {phase_str!r} in roadmap xlsx -- "
                f"expected one of {[p.value for p in RoadmapPhase]}"
            ) from exc

        ordinals[phase] += 1
        notes_raw = _safe_get(row, col_notes)
        sub_raw = _safe_get(row, col_sub)
        out.append({
            "phase": phase,
            "ordinal": ordinals[phase],
            "task_name": str(task_raw).strip(),
            "sub_tasks": _split_sub_tasks(sub_raw),
            "notes": str(notes_raw).strip() if not _looks_blank(notes_raw) else None,
        })
    return out


def import_roadmap_rows(
    session: Session,
    parsed: list[dict[str, Any]],
    *,
    overwrite: bool = False,
) -> dict[str, Any]:
    """Persist parsed rows into the ``roadmap_task`` table.

    Idempotency:
      - ``overwrite=False`` (default): refuses to import if any
        roadmap_task rows already exist.  Returns ``{"ok": False,
        "error": "..."}``.
      - ``overwrite=True``: DROPS all existing rows first, then inserts
        the new set.  This is the supported "re-import after editing
        the xlsx" path.

    Returns a summary dict counting rows per phase.  Does NOT commit --
    caller owns the transaction.
    """
    existing = session.query(RoadmapTask).count()
    if existing and not overwrite:
        return {
            "ok": False,
            "error": (
                f"roadmap_task already has {existing} rows.  Re-run "
                f"with --overwrite to replace them."
            ),
            "existing_count": existing,
        }
    if existing and overwrite:
        session.query(RoadmapTask).delete()
        session.flush()

    by_phase: dict[str, int] = {p.value: 0 for p in RoadmapPhase}
    for row in parsed:
        sub_tasks_json = (
            json.dumps(row["sub_tasks"]) if row["sub_tasks"] is not None else None
        )
        session.add(RoadmapTask(
            phase=row["phase"],
            ordinal=row["ordinal"],
            task_name=row["task_name"],
            sub_tasks_json=sub_tasks_json,
            notes=row.get("notes"),
        ))
        by_phase[row["phase"].value] += 1
    session.flush()

    return {
        "ok": True,
        "total": len(parsed),
        "by_phase": by_phase,
        "overwrote": existing,
    }


def list_roadmap_tasks(
    session: Session,
    *,
    actors: list[RoadmapActor] | None = None,
) -> list[dict[str, Any]]:
    """Return every roadmap task, ordered by phase then ordinal.

    JSON-serializable.  Used by the AI layer (Layer 2: prompt
    injection) and by any UI / report consumer.

    When ``actors`` is set, only tasks whose ``actor`` is in the list
    are returned -- this is how the proposal-prompt filter ignores
    architect-only tasks before injecting the roadmap into a
    contractor-side prompt.  Pass ``None`` (default) for all tasks.
    """
    from project_db.db.models import ROADMAP_PHASE_ORDER

    q = session.query(RoadmapTask)
    if actors is not None:
        q = q.filter(RoadmapTask.actor.in_(actors))
    rows = q.all()
    rows.sort(key=lambda r: (ROADMAP_PHASE_ORDER[r.phase], r.ordinal))
    out: list[dict[str, Any]] = []
    for r in rows:
        try:
            sub = json.loads(r.sub_tasks_json) if r.sub_tasks_json else None
        except (json.JSONDecodeError, TypeError):
            sub = None
        out.append({
            "canonical_id": str(r.canonical_id),
            "phase": r.phase.value if hasattr(r.phase, "value") else str(r.phase),
            "ordinal": int(r.ordinal),
            "task_name": r.task_name,
            "sub_tasks": sub,
            "notes": r.notes,
            "actor": (
                r.actor.value if r.actor is not None and hasattr(r.actor, "value")
                else (str(r.actor) if r.actor is not None else None)
            ),
        })
    return out


def classify_roadmap_actors(
    session: Session,
    provider,
) -> dict[str, Any]:
    """Use the LLM to draft an actor for each roadmap_task row.

    Sends the 44 tasks + sub-tasks in ONE call to Sonnet, asks for
    strict JSON ``{phase, ordinal, actor}`` per task.  Validates each
    item; bad ones go to errors, never crash the batch.  Writes the
    actor values back to ``roadmap_task``.  Does NOT commit -- caller
    owns the transaction.

    Returns ``{"ok": bool, "updated": int, "errors": list[str],
    "by_actor": {...}}``.
    """
    from project_db.ai.providers.base import LLMMessage

    tasks = list_roadmap_tasks(session)
    if not tasks:
        return {
            "ok": False,
            "error": "no roadmap_task rows -- run `import-roadmap` first.",
        }

    # Build the prompt.  Conservative posture; this is an extractor
    # not an analyst (per the askbot/proposal-bot prompt-philosophy
    # boundary).
    system = (
        "You classify construction project roadmap tasks by who is "
        "primarily responsible for executing them.  Three possible "
        "actors:\n"
        "- ARCHITECT: pure architect / designer work (e.g. site analysis, "
        "envelope detailing, code compliance drawings).\n"
        "- CONTRACTOR: pure contractor / builder work (e.g. cost "
        "estimating, fabricator coordination, punch list, "
        "preconstruction planning).\n"
        "- BOTH: genuinely co-responsible (e.g. project kickoff, "
        "client sign-off meetings, submittal review, clarification "
        "requests).\n\n"
        "Rules:\n"
        "- Use the task NAME and the SUB-TASKS as evidence.  When the "
        "sub-tasks describe drawing / modelling / certification work, "
        "lean ARCHITECT.  When they describe cost / RFP / build / "
        "installation work, lean CONTRACTOR.  When both are described "
        "or a sign-off / coordination action is named, BOTH.\n"
        "- Be conservative: prefer BOTH over guessing one side when "
        "the task could plausibly involve both.\n"
        "- Output STRICT JSON only.  No prose, no markdown fences."
    )

    lines = ["=== ROADMAP TASKS TO CLASSIFY ==="]
    for t in tasks:
        sub = " | ".join(t["sub_tasks"]) if t["sub_tasks"] else "(no sub-tasks)"
        lines.append(
            f'\n[{t["phase"]}-{t["ordinal"]}] "{t["task_name"]}"\n'
            f'  sub-tasks: {sub}'
        )

    user = (
        "\n".join(lines)
        + "\n\n---\n\n"
        "INSTRUCTION: Classify EVERY task above by actor.  Return strict "
        "JSON:\n\n"
        "{\n"
        '  "classifications": [\n'
        "    {\n"
        '      "phase": "SD" | "DD" | "CD" | "CA",\n'
        '      "ordinal": <int>,\n'
        '      "actor": "ARCHITECT" | "CONTRACTOR" | "BOTH",\n'
        '      "reasoning": "<one short sentence explaining the call>"\n'
        "    }, ...\n"
        "  ]\n"
        "}\n\n"
        f"Classify exactly {len(tasks)} task(s)."
    )

    try:
        raw = provider.complete_json(
            messages=[LLMMessage(role="user", content=user)],
            system=system,
            max_tokens=4000,
        )
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"LLM call failed: {exc}"}

    classifications = raw.get("classifications") or []
    if not isinstance(classifications, list):
        return {"ok": False, "error": "LLM returned no 'classifications' list"}

    # Index our rows by (phase, ordinal) for fast lookup.
    by_key: dict[tuple[str, int], RoadmapTask] = {}
    for r in session.query(RoadmapTask).all():
        phase_val = r.phase.value if hasattr(r.phase, "value") else str(r.phase)
        by_key[(phase_val, int(r.ordinal))] = r

    errors: list[str] = []
    by_actor: dict[str, int] = {a.value: 0 for a in RoadmapActor}
    updated = 0
    for item in classifications:
        if not isinstance(item, dict):
            errors.append(f"non-dict item: {item!r}")
            continue
        phase = str(item.get("phase", "")).upper()
        try:
            ordinal = int(item.get("ordinal"))
        except (TypeError, ValueError):
            errors.append(f"bad ordinal: {item!r}")
            continue
        actor_raw = str(item.get("actor", "")).upper()
        try:
            actor = RoadmapActor(actor_raw)
        except ValueError:
            errors.append(
                f"[{phase}-{ordinal}] unknown actor {actor_raw!r} "
                f"(expected one of {[a.value for a in RoadmapActor]})"
            )
            continue
        row = by_key.get((phase, ordinal))
        if row is None:
            errors.append(
                f"[{phase}-{ordinal}] no matching roadmap_task row"
            )
            continue
        row.actor = actor
        by_actor[actor.value] += 1
        updated += 1
    session.flush()

    return {
        "ok": True,
        "updated": updated,
        "by_actor": by_actor,
        "errors": errors,
        "total_classifications_returned": len(classifications),
    }
