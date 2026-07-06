"""Ingest a real, template-shaped SOW workbook into SowPackage / SowItem rows.

This is the first of the three "Drive go-live" ingesters (see HANDOFF): it turns
a `{code}_SOW_v1*.xlsx` file (the settled SOW template: a `SOW_Items` sheet with
Item_ID / CSI_Div_Code / Trade / Description / Included / Material_Spec / Notes)
into the structured scope spine, REPLACING any prior SOW rows for the project.

Built when the first REAL convention-shaped SOW file existed (Rockland,
2026-07-05) -- not before, so it is tested against real data, not speculation.

Deterministic, template-first (meeting plan §9: the deterministic path is
primary for templated SOP inputs; LLM tolerance is the fallback for legacy docs).
No LLM here.

Division handling mirrors the Phase-3 rule already in the schema:
  - one `SowPackage` per distinct CSI division EXCEPT `01` (General
    Requirements = GC overhead, no subcontractor package) -> its items get
    `package_id = NULL`;
  - division codes are canonicalized (`1012` -> `10-12`) so they match the
    resolver + budget/quote codes;
  - the sheet's own Trade label is kept as the package/item trade name (it is
    the project's real wording); the canonical CSI code is what joins.

Idempotent per project: existing SowPackage + SowItem rows for the project are
deleted and rebuilt on each call, so re-ingesting a revised SOW self-heals.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from project_db.ai.financial_divisions import canonical_division_code

# Divisions that are GC overhead, not a subcontractor trade package.
_NO_PACKAGE_DIVISIONS = {"01"}

_HEADERS = ("Item_ID", "CSI_Div_Code", "Trade", "Description", "Included", "Material_Spec", "Notes")


@dataclass
class SowIngestResult:
    project_id: object
    packages_created: int = 0
    items_created: int = 0
    included_items: int = 0
    excluded_items: int = 0
    divisions: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _rows_from_workbook(source) -> list[dict]:
    """Read the SOW_Items sheet into a list of dict rows (header-keyed).

    *source* is a path (str/Path) or raw bytes. Deterministic openpyxl read of
    the one structured sheet -- no evidence-bundle round-trip needed for a
    clean template. Raises ValueError if the SOW_Items sheet is missing.
    """
    import io

    import openpyxl

    if isinstance(source, (str, Path)):
        wb = openpyxl.load_workbook(Path(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)

    if "SOW_Items" not in wb.sheetnames:
        raise ValueError(f"workbook has no 'SOW_Items' sheet (sheets: {wb.sheetnames})")

    ws = wb["SOW_Items"]
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return []
    header = [str(c).strip() if c is not None else "" for c in raw[0]]
    out = []
    for r in raw[1:]:
        cells = list(r) + [None] * (len(header) - len(r))
        row = {header[i]: cells[i] for i in range(len(header))}
        item_id = row.get("Item_ID")
        if item_id is None or (isinstance(item_id, str) and not item_id.strip()):
            continue  # skip blank / total / separator rows
        out.append(row)
    return out


def ingest_sow_workbook(
    session, project, source, *, source_name: str | None = None
) -> SowIngestResult:
    """Replace *project*'s SOW spine from a template SOW workbook. See module docstring."""
    import json

    from project_db.db.models.sow import SowItem, SowPackage

    result = SowIngestResult(project_id=project.canonical_id)
    rows = _rows_from_workbook(source)
    if not rows:
        result.warnings.append("SOW_Items sheet had no data rows")
        return result

    # Idempotent: wipe this project's prior SOW rows (items before packages -- FK).
    session.query(SowItem).filter(SowItem.project_id == project.canonical_id).delete(
        synchronize_session="fetch"
    )
    session.query(SowPackage).filter(SowPackage.project_id == project.canonical_id).delete(
        synchronize_session="fetch"
    )
    session.flush()

    # Build one package per non-GC-overhead division (first trade label wins).
    packages: dict[str, object] = {}
    for row in rows:
        raw_div = row.get("CSI_Div_Code")
        div = canonical_division_code(str(raw_div).strip()) if raw_div is not None else "99"
        if div in _NO_PACKAGE_DIVISIONS or div in packages:
            continue
        trade = (str(row.get("Trade")).strip() if row.get("Trade") else None)
        pkg = SowPackage(
            project_id=project.canonical_id,
            division_code=div,
            trade_name=trade,
            title=f"{div}-{trade}" if trade else div,
            status="draft",
            source_meta_json=json.dumps({"source": source_name or "SOW workbook ingest"}),
        )
        session.add(pkg)
        packages[div] = pkg
    session.flush()

    seen_divs: set[str] = set()
    for row in rows:
        raw_div = row.get("CSI_Div_Code")
        div = canonical_division_code(str(raw_div).strip()) if raw_div is not None else "99"
        seen_divs.add(div)
        included_raw = str(row.get("Included") or "").strip().upper()
        included = included_raw != "N"  # anything not explicit "N" is included
        pkg = packages.get(div)
        spec_val = row.get("Material_Spec")
        material_spec = (str(spec_val).strip() or None) if spec_val else None
        item = SowItem(
            project_id=project.canonical_id,
            package_id=pkg.canonical_id if pkg is not None else None,
            item_code=(str(row.get("Item_ID")).strip() if row.get("Item_ID") else None),
            description=(str(row.get("Description")).strip() if row.get("Description") else None),
            division_code=div,
            included=included,
            material_spec=material_spec,
            source_meta_json=json.dumps(
                {
                    "source": source_name or "SOW workbook ingest",
                    "trade": (str(row.get("Trade")).strip() if row.get("Trade") else None),
                    "notes": (str(row.get("Notes")).strip() if row.get("Notes") else None),
                }
            ),
        )
        session.add(item)
        result.items_created += 1
        if included:
            result.included_items += 1
        else:
            result.excluded_items += 1

    session.flush()
    result.packages_created = len(packages)
    result.divisions = sorted(seen_divs)
    if "99" in seen_divs:
        result.warnings.append(
            "one or more SOW rows had a division code that did not resolve to a "
            "canonical CSI division (landed in 99) -- check CSI_Div_Code column"
        )
    return result
