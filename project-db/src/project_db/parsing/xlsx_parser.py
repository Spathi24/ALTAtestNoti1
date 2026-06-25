"""XLSX parser via openpyxl -- the highest-value parser of the refactor.

Spreadsheets were the worst troublemakers: flattening a multi-sheet workbook (or
two tables sitting side by side) into one text blob is exactly how a *supplier
cost worksheet* came to read like a client quote (3940's "Quotes"). This parser
PRESERVES structure: each sheet stays separate, the header row is detected, and
every value is reported as a ``{header: value}`` row -- so a downstream extractor
sees "Company = <vendor>, Quote = <amount>" and can tell cost from revenue.

Per sheet it emits one ``table_region`` EvidenceSpan carrying:
  - headers + a structured ``rows_sample`` (values associated with their column),
  - ``merged_ranges`` and a compact ``cells`` map of FORMULA cells (address ->
    raw/displayed/formula/number_format) so totals like ``=SUM(F4:F28)`` are
    citeable and distinguishable from line items,
  - a Markdown rendering for the ``DocumentText`` compatibility row.

Bounded (rows/cells/chars caps) so a giant model/projection workbook can't dump
megabytes. ``.xls`` (legacy BIFF) is NOT handled by openpyxl -> routes to a
``skipped`` parse. Pure: bytes in, ParsedDocument out, no DB/network.
"""

from __future__ import annotations

import io
import warnings

from project_db.parsing.base import ParsedDocument, ParsedEvidence
from project_db.parsing.tableutil import detect_header, is_formula

_XLSX_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel.sheet.macroenabled.12",  # .xlsm
}
_MAX_SHEETS = 50
_MAX_ROWS_PER_SHEET = 2000  # rows scanned for structure
_MAX_RENDER_ROWS = 200  # rows put into the compat Markdown
_MAX_SAMPLE_ROWS = 25  # rows put into rows_sample
_MAX_CELLS_MAP = 300  # formula/notable cells captured per sheet
_MAX_TOTAL_CHARS = 60_000  # whole-workbook rendered_text cap


def _a1(row: int, col: int) -> str:
    from openpyxl.utils import get_column_letter

    return f"{get_column_letter(col)}{row}"


def _cell(x: object) -> str:
    s = "" if x is None else str(x)
    return s.replace("|", "\\|").replace("\r", " ").replace("\n", " ").strip()


def _row_to_md(cells: list, width: int) -> str:
    rr = list(cells) + [""] * (width - len(cells))
    return "| " + " | ".join(_cell(c) for c in rr[:width]) + " |"


class XlsxParser:
    name = "xlsx"
    version = "1"

    def can_parse(self, *, mime: str | None, filename: str | None) -> bool:
        if mime and mime.lower().split(";")[0].strip() in _XLSX_MIMES:
            return True
        if filename and filename.lower().endswith((".xlsx", ".xlsm")):
            return True
        return False

    def parse(self, content: bytes, *, doc_name: str, mime: str | None) -> ParsedDocument:
        from openpyxl import load_workbook

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            # data_only=False -> formula cells return their formula string; plain
            # cells return their value. We pull cached/displayed values lazily for
            # the (usually few) formula cells via a second targeted load.
            wb = load_workbook(io.BytesIO(content), read_only=False, data_only=False)

        sheet_names = list(wb.sheetnames)
        spans: list[ParsedEvidence] = []
        sheet_meta: list[dict] = []
        render_parts: list[str] = []
        total_chars = 0
        cached_wb = None  # lazily loaded data_only workbook for formula results

        try:
            for ws in wb.worksheets[:_MAX_SHEETS]:
                max_row = min(ws.max_row or 0, _MAX_ROWS_PER_SHEET)
                max_col = ws.max_column or 0
                merged = (
                    [str(r) for r in getattr(ws, "merged_cells", []).ranges][:50] if max_col else []
                )

                rows: list[list] = []
                formula_cells: dict[str, dict] = {}
                for row in ws.iter_rows(
                    min_row=1, max_row=max_row, max_col=max_col, values_only=False
                ):
                    values = [c.value for c in row]
                    if not any(v not in (None, "") for v in values):
                        continue
                    rows.append([c.value for c in row])
                    for c in row:
                        if is_formula(c.value) and len(formula_cells) < _MAX_CELLS_MAP:
                            formula_cells[c.coordinate] = {
                                "formula": c.value,
                                "number_format": c.number_format,
                                "raw_value": None,
                                "displayed_value": None,
                            }

                if not rows:
                    sheet_meta.append(
                        {
                            "name": ws.title,
                            "max_row": 0,
                            "max_col": max_col,
                            "n_merged": len(merged),
                        }
                    )
                    continue

                # Resolve formula results from a data_only load, only if needed.
                if formula_cells:
                    if cached_wb is None:
                        with warnings.catch_warnings():
                            warnings.simplefilter("ignore")
                            cached_wb = load_workbook(
                                io.BytesIO(content), read_only=False, data_only=True
                            )
                    cws = cached_wb[ws.title]
                    for addr in formula_cells:
                        try:
                            formula_cells[addr]["displayed_value"] = cws[addr].value
                        except Exception:
                            pass

                header_idx, header_conf = detect_header(rows)
                headers = [(_cell(h) or f"col{i}") for i, h in enumerate(rows[header_idx])]
                width = max(len(headers), max((len(r) for r in rows), default=0))
                data_rows = rows[header_idx + 1 :]
                sample = [
                    {
                        (headers[i] if i < len(headers) else f"col{i}"): (
                            v if not is_formula(v) else str(v)
                        )
                        for i, v in enumerate(r)
                        if v not in (None, "")
                    }
                    for r in data_rows[:_MAX_SAMPLE_ROWS]
                ]
                # Raw grid safety net: never lose structure even if header
                # detection is imperfect (title rows, multi-table sheets).
                rows_preview = [
                    [None if v is None else (str(v) if is_formula(v) else v) for v in r]
                    for r in rows[: _MAX_SAMPLE_ROWS + header_idx + 1]
                ]

                md_lines = [f"## Sheet: {ws.title}"]
                # Preserve any title/metadata rows above the detected header.
                for r in rows[:header_idx]:
                    md_lines.append(_row_to_md(r, width))
                md_lines.append(_row_to_md(headers, width))
                md_lines.append("| " + " | ".join("---" for _ in range(width)) + " |")
                for r in data_rows[:_MAX_RENDER_ROWS]:
                    md_lines.append(_row_to_md(r, width))
                if len(data_rows) > _MAX_RENDER_ROWS:
                    md_lines.append(f"| ...({len(data_rows) - _MAX_RENDER_ROWS} more rows) |")
                sheet_md = "\n".join(md_lines)

                a1_range = f"A1:{_a1(len(rows), max(width, 1))}"
                spans.append(
                    ParsedEvidence(
                        evidence_type="table_region",
                        locator={
                            "sheet": ws.title,
                            "range": a1_range,
                            "header_row": header_idx + 1,
                            "n_rows": len(data_rows),
                            "n_cols": width,
                        },
                        content_text=sheet_md[:4000],
                        content_json={
                            "sheet": ws.title,
                            "headers": headers,
                            "n_rows": len(data_rows),
                            "n_cols": width,
                            "merged_ranges": merged,
                            "rows_sample": sample,
                            "rows_preview": rows_preview,
                            "cells": formula_cells,
                            "header_confidence": header_conf,
                        },
                        confidence=header_conf,
                    )
                )
                sheet_meta.append(
                    {
                        "name": ws.title,
                        "max_row": len(rows),
                        "max_col": width,
                        "n_merged": len(merged),
                        "n_formulas": len(formula_cells),
                    }
                )
                if total_chars < _MAX_TOTAL_CHARS:
                    render_parts.append(sheet_md)
                    total_chars += len(sheet_md) + 1
        finally:
            wb.close()
            if cached_wb is not None:
                cached_wb.close()

        rendered_text = "\n\n".join(render_parts).strip()
        structured = {
            "format": "xlsx",
            "n_sheets": len(sheet_names),
            "sheet_names": sheet_names,
            "sheets": sheet_meta,
        }
        return ParsedDocument(
            rendered_text=rendered_text, structured=structured, evidence_spans=spans
        )
